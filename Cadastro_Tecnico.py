from tkinter import*
from tkinter import ttk
from random import randint
from turtle import width
from openpyxl import Workbook, load_workbook



path = ('cad_tecnico.xlsx')
cad_excel = load_workbook(path)
lista_cad = []



janela_tecnico = Tk()

janela_tecnico.title('.:. Cadastro de Técnicos .:.')
janela_tecnico.geometry("800x500")
 

def inserir_tecnico():
    nome = entry_nome.get()
    cpf = entry_cpf.get()
    telefone = entry_telefone.get()
    equipe = entry_equipe.get()
    lista_cad.append(nome)
    lista_cad.append(cpf)
    lista_cad.append(telefone)
    lista_cad.append(equipe)
 
def gravar_list():     
    plan_tec = cad_excel.active
    plan_tec.append(lista_cad)
    cad_excel.close()
    cad_excel.save(path)


def alterar_list():
    for d in cad_excel['cadastro'].iter_rows(values_only = True):        
     print(d)



def format_cpf(event = None):
    
    text = entry_cpf.get().replace(".", "").replace("-", "")[:11]
    new_text = ""

    if event.keysym.lower() == "backspace": return
    
    for index in range(len(text)):
        
        if not text[index] in "0123456789": continue
        if index in [2, 5]: new_text += text[index] + "."
        elif index == 8: new_text += text[index] + "-"
        else: new_text += text[index]

    entry_cpf.delete(0, "end")
    entry_cpf.insert(0, new_text)


#Rótulos Entradas...
label_nome = Label(janela_tecnico, text='Nome')
label_nome.grid(row=0,column=0, padx=10, pady=10)

label_cpf = Label(janela_tecnico, text='CPF (Digite somente números)')
label_cpf.grid(row=1, column=0, padx=10, pady=10)

label_telefone = Label(janela_tecnico, text='Telefone')
label_telefone.grid(row=2, column=0 , padx=10, pady=10)

label_equipe = Label(janela_tecnico, text='Equipe')
label_equipe.grid(row=3, column=0, padx=10, pady=10)

#Caixas Entradas...
entry_nome = Entry(janela_tecnico , width =35)
entry_nome.grid(row=0,column=1, padx=10, pady=10)

entry_cpf = Entry(janela_tecnico, width =35)
entry_cpf.grid(row=1, column=1, padx=10, pady=10)
entry_cpf.bind("<KeyRelease>", format_cpf) # Formatando o CPF.

entry_telefone = Entry(janela_tecnico, width =35)
entry_telefone.grid(row=2, column=1 , padx=10, pady=10)

entry_equipe = Entry(janela_tecnico, width =35)
entry_equipe.grid(row=3, column=1, padx=10, pady=10)

# Gerando Botões...
botao_cadastrar = Button(text='Inserir',command=inserir_tecnico)
botao_cadastrar.grid(row=0, column=3,columnspan=1, padx=10, pady=10 , ipadx = 30)

botao_salvar = Button(text='Salvar',command=gravar_list)
botao_salvar.grid(row=1, column=3,columnspan=1, padx=10, pady=10 , ipadx = 30)

botao_alterar = Button(text='Alterar',command=alterar_list)
botao_alterar.grid(row=2, column=3,columnspan=1, padx=10, pady=10 , ipadx = 30)

botao_deletar = Button(text='Deletar',command=janela_tecnico.quit)
botao_deletar.grid(row=3, column=3,columnspan=1, padx=10, pady=10 , ipadx = 30)

botao_deletar = Button(text='Atualizar',command=janela_tecnico.quit)
botao_deletar.grid(row=4, column=3,columnspan=1, padx=10, pady=10 , ipadx = 30)

# Gerando Treeview...

fbd = Frame(width=706,height=496,bg = '#4169E1',relief='flat')
fbd.grid(column=0,row=7)
tv = ttk.Treeview(fbd,columns=('Nome','CPF','Telefone','Equipe'),show='headings')


tv.column('Nome',minwidth=0,width=70)
tv.column('CPF',minwidth=0,width=124)
tv.column('Telefone',minwidth=0,width=70)
tv.column('Equipe',minwidth=0,width=148)
tv.heading('Nome',text='Nome')
tv.heading('CPF',text='CPF')
tv.heading('Telefone',text='Telefone')
tv.heading('Equipe',text='Equipe')
tv.pack()


janela_tecnico.mainloop()




























