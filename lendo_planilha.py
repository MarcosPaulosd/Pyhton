import openpyxl
import datetime
from tkinter import *
from tkinter import ttk
import os

class lendo:


    def planilha_tec(tv):
        plt = openpyxl.load_workbook('controle_ferramenta.xlsx')['tecnico']
        for x in plt:
            id = x[0].value
            nome = x[1].value
            cpf = x[2].value
            telefone = x[3].value
            turno = x[4].value
            equipe = x[5].value

            tv.insert("","end",values = (id,nome,cpf,telefone,turno,equipe))



    def lendo_planilha(tv):
        pl = openpyxl.load_workbook('controle_ferramenta.xlsx')['ferramenta']

        for x in pl:
            if x[11].value == 'indisponivel':
                
                id = x[0].value
                ferramenta = x[1].value
                idtec = x[12].value
                nometec = x[13].value
                telefonetec = x[14].value
                dataentrega = x[15].value

                tv.insert("","end",values = (id,ferramenta,idtec,nometec,telefonetec,dataentrega))


    def comboboxtecnico():
        listatec = ['Selecione o Tecnico     -- > ']
        pl = openpyxl.load_workbook('controle_ferramenta.xlsx')['tecnico']
        for x in pl:
            t = f'{x[0].value} {x[1].value}'
            listatec.append(t)
        return listatec


    def comboboxferramenta():
        listaferr = ['Selecione a ferramenta     -- > ']
        pl = openpyxl.load_workbook('controle_ferramenta.xlsx')['ferramenta']
        for x in pl:
            t = f'{x[0].value} {x[1].value}'
            listaferr.append(t)
        return listaferr


    def reservando(lista):
        listatec = []
        listafer = []
        
        pltec = openpyxl.load_workbook('controle_ferramenta.xlsx')['tecnico']

        for x in pltec:
            if str(x[0].value) == lista[0]:
                listatec.extend([x[0].value,x[1].value,x[3].value])

        plfer = openpyxl.load_workbook('controle_ferramenta.xlsx')
        plfercell = plfer['ferramenta']

        for x in plfercell:
            if str(x[0].value) == lista[1]:
                x[11].value = 'indisponivel'
                x[12].value = listatec[0]

                x[13].value = listatec[1]
                x[14].value = listatec[2]
                x[15].value = lista[3]

                listafer.extend([x[0].value,x[1].value,listatec[0],listatec[1],lista[2],lista[3]])
                print(listafer)

        
        plfer.save('controle_ferramenta.xlsx')


        plhis = openpyxl.load_workbook('controle_ferramenta.xlsx')
        plhiscell = plhis['historico']

        plhiscell.append(listafer)

        plhis.save('controle_ferramenta.xlsx')





    def lendo_ferramentas(tv):
        pl = openpyxl.load_workbook('controle_ferramenta.xlsx')['ferramenta']

        for x in pl:
            id = x[0].value
            ferramenta = x[1].value
            numeroserial = x[3].value
            voltagem = x[2].value
            tipo = x[7].value

            tv.insert("","end",values=(id,ferramenta,numeroserial,voltagem,tipo))


    def combobox_devolucao():
        lista = ['Selecione a ferramenta']
        pl = openpyxl.load_workbook('controle_ferramenta.xlsx')['ferramenta']

        for x in pl:
            if x[11].value == 'indisponivel':
                a = f'{x[0].value} {x[1].value}'
                lista.append(a)

        return lista


    def devolvendo(lista,tv):
        listadev = []
        hoje = datetime.datetime.today().strftime('%d/%m/%y')


        idultimotec = ''
        ultimotec = ''
        
        idinteiro = lista[0]
        e = idinteiro.find(' ')
        idseparado = idinteiro[:e]
        ferramenta = idinteiro[e:]
        




        ph = openpyxl.load_workbook('controle_ferramenta.xlsx')
        phcell = ph['historico']

        for x in phcell:
            if str(x[0].value) == idseparado:
                idultimotec = x[2].value
                ultimotec = x[3].value

        listadev.extend([idseparado,ferramenta,idultimotec,ultimotec,lista[1],hoje])
        print(listadev)


        phcell.append(listadev)

        ph.save('controle_ferramenta.xlsx')

        phfer = openpyxl.load_workbook('controle_ferramenta.xlsx')
        phfercell = phfer['ferramenta']

        for x in phfercell:
            if str(x[0].value) == idseparado:
                x[11].value = 'disponivel'
                x[12].value = ''
                x[13].value = ''
                x[14].value = ''
                x[15].value = ''

        
        phfer.save('controle_ferramenta.xlsx')


        for item in tv.get_children():
            tv.delete(item)

        lendo.lendo_planilha(tv)
        return tv



    def extraindo_ferramentas():
        plf = openpyxl.load_workbook('controle_ferramenta.xlsx')['ferramenta']

        planilhanova = openpyxl.Workbook()
        cell = planilhanova['Sheet']
        cabec = ['ID','Fabricante','Voltagem','Numero de série','altura','largura','profundidade','Tipo','Material','Tempo de uso máx','Descrição']
        cell.append(cabec)


        for x in plf:
            listafe = []
            id= x[0].value
            fe = x[1].value
            voltagem = x[2].value
            ns = x[3].value
            al = f'{x[4].value}cm'
            lar = f'{x[5].value}cm'
            pro = f'{x[6].value}cm'
            tipo = x[7].value
            mat = x[8].value
            tempo = f'{x[9].value} dias'
            desc = x[10].value

            listafe.extend([id,fe,voltagem,ns,al,lar,pro,tipo,mat,tempo,desc])
            

            cell.append(listafe)

        planilhanova.save('planilha_ferramenta.xlsx')

        source = "planilha_ferramenta.xlsx"
        destination = "C:\\Users\\Henrique\\Desktop\\destino\\planilha_ferramenta.xlsx"

        try:
            if os.path.exists(destination):
                print('esse item já existe')
                os.remove("C:\\Users\\Henrique\\Desktop\\destino\\planilha_ferramenta.xlsx")
                os.replace(source,destination)
                print("deletado e alterado")


            else:
                os.replace(source,destination)
                print('tudo certo')

        except FileNotFoundError:
            print(source + ' was not found')


    def extraindo_tecnicos():

        plt = openpyxl.load_workbook('controle_ferramenta.xlsx')['tecnico']

        planilhanova = openpyxl.Workbook()
        cell = planilhanova['Sheet']
        cabec = ['ID','Nome','CPF','Telefone','Turno','Equipe']
        cell.append(cabec)

        for x in plt:
            listatec =[]
            id = x[0].value
            nome = x[1].value
            cpf = x[2].value
            telefone = x[3].value
            turno = x[4].value
            equipe = x[5].value
            listatec.extend([id,nome,cpf,telefone,turno,equipe])

            cell.append(listatec)


        planilhanova.save('planilha_tecnico.xlsx')

        source = "planilha_tecnico.xlsx"
        destination = "C:\\Users\\Henrique\\Desktop\\destino\\planilha_tecnico.xlsx"

        try:
            if os.path.exists(destination):
                print('esse item já existe')
                os.remove("C:\\Users\\Henrique\\Desktop\\destino\\planilha_tecnico.xlsx")
                os.replace(source,destination)
                print("deletado e alterado")


            else:
                os.replace(source,destination)
                print('tudo certo')

        except FileNotFoundError:
            print(source + ' was not found')



    def extraindo_historico():
        plh = openpyxl.load_workbook('controle_ferramenta.xlsx')['historico']

        planilhanova = openpyxl.Workbook()

        cell = planilhanova['Sheet']
        cabec = ['Id ferramenta','Fabricante','Id Tecnico','Nome Tecnico','Data de busca','Data de entrega']
        cell.append(cabec)


        for x in plh:
            listahis = []

            idfer = x[0].value
            fer = x[1].value
            idtec = x[2].value
            nometec = x[3].value
            datab = x[4].value
            datae = x[5].value

            listahis.extend([idfer,fer,idtec,nometec,datab,datae])

            cell.append(listahis)


        planilhanova.save('planilha_historico.xlsx')

        source = "planilha_historico.xlsx"
        destination = "C:\\Users\\Henrique\\Desktop\\destino\\planilha_historico.xlsx"

        try:
            if os.path.exists(destination):
                print('esse item já existe')
                os.remove("C:\\Users\\Henrique\\Desktop\\destino\\planilha_historico.xlsx")
                os.replace(source,destination)
                print("deletado e alterado")


            else:
                os.replace(source,destination)
                print('tudo certo')

        except FileNotFoundError:
            print(source + ' was not found')

