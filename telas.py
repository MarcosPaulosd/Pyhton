from ftplib import error_reply
from tkinter import *
from tkinter import ttk
import openpyxl
from lendo_planilha import lendo

class telas_cadastro:

    def jan_cadastro_ferramenta():


        tl2 = Toplevel()


        tl2.title('Cadastro de Ferramenta')
        tl2.geometry('1200x400')
        tl2.configure(background='#35363a')

        Label(tl2,text='Fabrincante',bg = '#35363a',fg='#FFFFFF',font=('Arial 12 bold')).place(x=10,y=10)
        efabricante = Entry(tl2)
        efabricante.place(x=10,y=30)
        errofabricante = Label(tl2,text='',bg = '#35363a',fg='#FF0000')
        errofabricante.place(x=8,y=50)
        
        Label(tl2,text='Voltagem de uso',bg = '#35363a',fg='#FFFFFF',font=('Arial 12 bold')).place(x=180,y=10)
        voltagens = ['Selecione a voltagem','Não usa energia','110v','220v']
        evoltagem = ttk.Combobox(tl2,values=voltagens)
        evoltagem.set('Selecione a voltagem')
        evoltagem.place(x=180,y=30)
        errovoltagem = Label(tl2,text='',bg = '#35363a',fg='#FF0000')
        errovoltagem.place(x=178,y=50)

        Label(tl2,text='Numero de serie',bg = '#35363a',fg='#FFFFFF',font=('Arial 12 bold')).place(x=350,y=10)
        ens = Entry(tl2,width=40)
        ens.place(x = 350,y=30)
        erroserial = Label(tl2,text='',bg = '#35363a',fg='#FF0000')
        erroserial.place(x=348,y=50)

        Label(tl2,text='Tamanho ( em cm )',bg = '#35363a',fg='#FFFFFF',font=('Arial 12 bold')).place(x=10,y=70)
        Label(tl2,text='Altura',bg = '#35363a',fg='#FFFFFF').place(x=19,y=95)
        ealtura = Entry(tl2,width=10)
        ealtura.place(x=57,y =97)
        Label(tl2,text='Largura',bg = '#35363a',fg='#FFFFFF').place(x=125,y =95)
        elargura = Entry(tl2,width=10)
        elargura.place(x=175,y=97)
        Label(tl2,text='Profundidade',bg = '#35363a',fg='#FFFFFF').place(x=245,y=95)
        eprofundidade = Entry(tl2,width=10)
        eprofundidade.place(x=325,y=97)
        errotamanho = Label(tl2,text='',bg = '#35363a',fg='#FF0000')
        errotamanho.place(x=150,y=120)

        Label(tl2,text='Tipo de Ferramenta',bg = '#35363a',fg='#FFFFFF',font=('Arial 12 bold')).place(x = 420,y=70)
        etipo = Entry(tl2,width=28)
        etipo.place(x=422,y=97)
        errotipo = Label(tl2,text='',bg = '#35363a',fg='#FF0000')
        errotipo.place(x=422,y=120)

        Label(tl2,text='Material da ferramenta',bg = '#35363a',fg='#FFFFFF',font=('Arial 12 bold')).place(x= 10,y = 140)
        ematerial = Entry(tl2,width=30)
        ematerial.place(x=10,y=160)
        erromaterial = Label(tl2,text='',bg = '#35363a',fg='#FF0000')
        erromaterial.place(x=10,y=180)

        Label(tl2,text='Tempo máximo de uso ( em dias )',bg = '#35363a',fg='#FFFFFF',font=('Arial 12 bold')).place(x=250,y=140)
        etempo = Entry(tl2,width=20)
        etempo.place(x=252,y=160)
        errotempo = Label(tl2,text='',bg = '#35363a',fg='#FF0000')
        errotempo.place(x=252,y=180)

        Label(tl2,text='Descrição da Ferramenta',bg = '#35363a',fg='#FFFFFF',font=('Arial 12 bold')).place(x=10,y=200)
        edesc = Text(tl2,width=30,height=3)
        edesc.place(x=10,y=220)
        errodesc = Label(tl2,text='',bg = '#35363a',fg='#FF0000')
        errodesc.place(x=10,y=280)

        cadastro_concluido = Label(tl2,text='',bg = '#35363a',fg='#33e327')
        cadastro_concluido.place(x= 160,y=330)



        tv = ttk.Treeview(tl2,height=18,columns=('id','ferramenta','numeroserial','voltagem','tipo'),show='headings')

        tv.column('id',minwidth=0,width=70)
        tv.column('ferramenta',minwidth=0,width=124)
        tv.column('numeroserial',minwidth=0,width=70)
        tv.column('voltagem',minwidth=0,width=148)
        tv.column('tipo',minwidth=0,width=148)

        tv.heading('id',text='ID ferrmanta')
        tv.heading('ferramenta',text='Fabricante')
        tv.heading('numeroserial',text='Numero de série')
        tv.heading('voltagem',text='Voltagem')
        tv.heading('tipo',text='Tipo')

        tv.place(x=600,y= 0)

        lendo.lendo_ferramentas(tv)

          


        def cadastrando_ferramenta():
            erro = False
            lista = []


            def identificando_id():
                pl = openpyxl.load_workbook('controle_ferramenta.xlsx')['ferramenta']
                id = 1
                for x in pl:
                    id = x[0].value
                id += 1
                return id
        
            id = identificando_id()
            lista.append(id)

            fabricante = efabricante.get()
            if fabricante == '' or len(fabricante) > 30:
                erro = True
                errofabricante['text'] = 'Fabricante error'
            else:
                errofabricante['text'] = ''
                lista.append(fabricante)
            

            voltagem = evoltagem.get()
            if voltagem == 'Selecione a voltagem':
                erro = True
                errovoltagem['text'] = 'Voltagem Error'
            else:
                errovoltagem['text'] = ''
                lista.append(voltagem)


            numeroserial = ens.get()
            if numeroserial == '' or len(numeroserial) > 25:
                erro = True
                erroserial['text'] = 'Numero de serie error'
            else:
                erroserial['text'] = ''
                lista.append(numeroserial)

            altura = ealtura.get()
            if altura == '':
                erro = True
                errotamanho['text'] = 'Tamanho Error'
            else:
                errotamanho['text'] = ''
                lista.append(altura)

            largura = elargura.get()
            if largura == '':
                erro = True
                errotamanho['text'] = 'Tamanho Error'
            else:
                errotamanho['text'] = ''
                lista.append(largura)

            profundidade = eprofundidade.get()
            if profundidade == '':
                erro = True
                errotamanho['text'] = 'Tamanho Error'
            else:
                errotamanho['text'] = ''
                lista.append(profundidade)

            tipo = etipo.get()
            if tipo == '' or len(tipo) > 15:
                erro = True
                errotipo['text'] = 'Tipo Error'
            else:
                errotipo['text'] = ''
                lista.append(tipo)

            material = ematerial.get()
            if material == '' or len(material) > 15:
                erro = True
                erromaterial['text'] = 'Material Error'
            else:
                erromaterial['text'] = ''
                lista.append(material)

            tempo = etempo.get()
            if tempo == '':
                erro = True
                errotempo['text'] = 'Tempo Error'
            else:
                errotempo['text'] = ''
                lista.append(tempo)

            descricao = edesc.get("1.0",END)
            if descricao == '\n' or len(descricao) > 60:
                erro = True
                errodesc['text'] = 'Descricao Error'
            else:
                errodesc['text'] = ''
                lista.append(descricao)

            if erro == False:
                lista.extend(['disponivel','none','none','none','none'])

                pl = openpyxl.load_workbook('controle_ferramenta.xlsx')
                plf = pl['ferramenta']
                plf.append(lista)
                pl.save('controle_ferramenta.xlsx')

                cadastro_concluido['text'] = 'Cadastro Concluido'

                efabricante.delete(0,END)
                evoltagem.set('Selecione a voltagem')
                ens.delete(0,END)
                ealtura.delete(0,END)
                elargura.delete(0,END)
                eprofundidade.delete(0,END)
                etipo.delete(0,END)
                ematerial.delete(0,END)
                etempo.delete(0,END)
                edesc.delete("1.0",END)

                for item in tv.get_children():
                    tv.delete(item)

                lendo.lendo_ferramentas(tv)
           
                       

        botao_cadastro = Button(tl2,text='Cadastrar',width=10,command=cadastrando_ferramenta)
        botao_cadastro.place(x=180,y=360)

        botao_fechar = Button(tl2,text='Fechar',width=10,command=tl2.destroy)
        botao_fechar.place(x= 330,y=360)


    def janela_cadastro_tecnico():

        tl3 = Toplevel()
        tl3.title('.:. Cadastro de Técnicos .:.')
        tl3.geometry("900x500")
      
        def format_cpf(event = None):
    
            text = entry_cpf.get().replace(".", "").replace("-", "")[:11]
            new_text = ""

            if event.keysym.lower() == "backspace": return
            
            for index in range(len(text)):
                
                if not text[index] in "0123456789": continue
                if index in [2, 5]: new_text += text[index] + "."
                elif index == 8: new_text += text[index] + "-"
                else: new_text += text[index]

            entry_cpf.delete(0, "end")
            entry_cpf.insert(0, new_text)


        def cadastrar_tecnico():
            def gerando_id():
                id = 0
                pl = openpyxl.load_workbook('controle_ferramenta.xlsx')
                plt = pl['tecnico']

                for x in plt:
                    id = x[0].value

                id += 1
                return id

            


            id = gerando_id()
            nome = entry_nome.get()
            cpf = entry_cpf.get()
            tel = entry_telefone.get()
            turno = entry_turno.get()
            equipe = entry_equipe.get()
            lista = [id,nome,cpf,tel,turno,equipe]


            errorcpf = False

            pl = openpyxl.load_workbook('controle_ferramenta.xlsx')
            plt = pl['tecnico']

            for x in plt:
                if x[2].value == cpf:
                    errorcpf = True
            

            if errorcpf == False:
                plt.append(lista)
                errorcadastro['text'] = 'Cadastro concluido'
                entry_cpf.delete(0,END)
                entry_nome.delete(0,END)
                entry_telefone.delete(0,END)
                entry_equipe.delete(0,END)
                entry_turno.set('Selecione turno')


                

                





            else:
                errorcadastro['text'] = 'Cadastro error'

            
            pl.save('controle_ferramenta.xlsx')

            for item in tv.get_children():
                    tv.delete(item)

            lendo.planilha_tec(tv)

            

            
        
        #Rótulos Entradas...
        label_nome = Label(tl3, text='Nome')
        label_nome.grid(row=0,column=0, padx=10, pady=10)

        label_cpf = Label(tl3, text='CPF (Digite somente números)')
        label_cpf.grid(row=1, column=0, padx=10, pady=10)

        label_telefone = Label(tl3, text='Telefone')
        label_telefone.grid(row=2, column=0 , padx=10, pady=10)

        label_equipe = Label(tl3, text='Equipe')
        label_equipe.grid(row=3, column=0, padx=10, pady=10)

        label_turno = Label(tl3, text='Equipe')
        label_turno.grid(row=4, column=0, padx=10, pady=10)

        #Caixas Entradas...
        entry_nome = Entry(tl3, width =35)
        entry_nome.grid(row=0,column=1, padx=10, pady=10)

        entry_cpf = Entry(tl3, width =35)
        entry_cpf.grid(row=1, column=1, padx=10, pady=10)
        entry_cpf.bind("<KeyRelease>", format_cpf) # Formatando o CPF.

        entry_telefone = Entry(tl3, width =35)
        entry_telefone.grid(row=2, column=1 , padx=10, pady=10)

        entry_equipe = Entry(tl3, width =35)
        entry_equipe.grid(row=3, column=1, padx=10, pady=10)

        listturno = ['Selecione turno','manhã','tarde','noite']
        entry_turno = ttk.Combobox(tl3,values=listturno)
        entry_turno.set('Selecione turno')
        entry_turno.grid(row=4,column=1)


        # Gerando Botões...

        botao_salvar = Button(tl3,text='Cadastrar',command=cadastrar_tecnico)
        botao_salvar.grid(row=1, column=3,columnspan=1, padx=10, pady=10 , ipadx = 30)

        # Gerando Treeview...

        fbd = Frame(tl3,width=706,height=496,bg = '#4169E1',relief='flat')
        fbd.grid(column=0,row=7)
        tv = ttk.Treeview(fbd,columns=('Id','Nome','CPF','Telefone','turno','Equipe'),show='headings')

        tv.column('Id',minwidth=0,width=30)
        tv.column('Nome',minwidth=0,width=70)
        tv.column('CPF',minwidth=0,width=124)
        tv.column('Telefone',minwidth=0,width=70)
        tv.column('turno',minwidth=0,width=60)
        tv.column('Equipe',minwidth=0,width=90)

        tv.heading('Id',text='ID')
        tv.heading('Nome',text='Nome')
        tv.heading('CPF',text='CPF')
        tv.heading('Telefone',text='Telefone')
        tv.heading('turno',text='Turno')
        tv.heading('Equipe',text='Equipe')
        tv.grid(column=0,row=0)

        lendo.planilha_tec(tv)


        botao_fechar = Button(tl3,text='Fechar',width=10,command=tl3.destroy)
        botao_fechar.grid(row=2,column=3)


        errorcadastro = Label(tl3, text='')
        errorcadastro.grid(row=3,column=3)
        

